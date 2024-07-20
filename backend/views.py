from django.shortcuts import render,HttpResponse,redirect
from django.contrib.auth.mixins import LoginRequiredMixin
from .forms import *
from django.views import View
from django.views.generic.list import ListView
from .models import *
from django.contrib.auth import logout
from num2words import num2words
import datetime
import openpyxl as xl;


# Create your views here.

def validator(orders):
    truthvalue = False
    for i in orders:
        order = OrderForm(i.cleaned_data)
        if(order.is_valid()):
            truthvalue = True
        
    return truthvalue

def Xlgenrator(context):

    invoice_no = context['invoice'].id
    buyer_address = context['buyer'].address
    orders = len(context['orders'])
    total_psc = context['total_pices']
    total_amo = context['total_amount']
    gst = float(context['cs_gst'])
    grand_total_amo = context['grand_total']
    word_total_amount = context['word_grand_total']
    word_tax_total = context['words_tax_total']

    positions = {'5': '', '17': '', '18': '', '28': '', '31': '', '32': '', '34': '', '36': '', '39': '', '40': '', '42': '', '44': '', '47': '', '48': '', '50': '', '52': '', '55': '', '56': '', '58': '', '60': '', '61': '', '66': '', '69': '', '72': '', '73': '', '74': '', '87': '', '89': '', '91': '', '93': '', '94': '', '95': '', '96': ''}

    if(orders == 4):
        p4 = context['orders'][0]
        singlep4 = p4.price_per_PSC
        q4 = p4.quantity
        amo4 = q4 * singlep4

        p3 = context['orders'][1]
        q3 = p3.quantity
        singlep3 = p3.price_per_PSC
        amo3 = q3 * singlep3
        
        p2 = context['orders'][2]
        singlep2 = p2.price_per_PSC
        q2 = p2.quantity
        amo2 = q2 * singlep2

        p1 = context['orders'][3]
        q1 = p1.quantity
        singlep1 = p1.price_per_PSC
        amo1 = q1 * singlep1
    
        positions = {'5': invoice_no, '17': buyer_address, '18': buyer_address, '28': p1.product.description, '31': q1, '32': singlep1, '34': amo1, '36': p2.product.description, '39': q2, '40': singlep2, '42': amo2, '44': p3.product.description, '47': q3, '48': singlep3, '50': amo3, '52': p4.product.description, '55': q4, '56': singlep4, '58': amo4, '60': total_psc, '61': total_amo, '66': gst, '69': gst, '72': total_psc, '73': grand_total_amo, '74': word_total_amount, '87':total_amo, '89': gst, '91': gst, '93': gst, '94': gst, '95': gst*2, '96': word_tax_total}

    elif(orders == 3):

        p1 = context['orders'][0]
        singlep1 = p1.price_per_PSC
        q1 = p1.quantity
        amo1 = q1 * singlep1

        p2 = context['orders'][1]
        singlep2 = p2.price_per_PSC
        amo2 = q2 * singlep2
        q2 = p2.quantity
        
        p3 = context['orders'][2]
        singlep3 = p3.price_per_PSC
        q3 = p3.quantity
        amo3 = q3 * singlep3

        positions = {'5': invoice_no, '17': buyer_address, '18': buyer_address, '28': p1.product.description, '31': q1, '32': singlep1, '34': amo1, '36': p2.product.description, '39': q2, '40': singlep2, '42': amo2, '44': p3.product.description, '47': q3, '48': singlep3, '50': amo3, '52': '', '55': '', '56': '', '58': '', '60': total_psc, '61': total_amo, '66': gst, '69': gst, '72': total_psc, '73': grand_total_amo, '74': word_total_amount, '87':total_amo, '89': gst, '91': gst, '93': gst, '94': gst, '95': gst*2, '96': word_tax_total}

    elif(orders == 2):

        p1 = context['orders'][0]
        singlep1 = p1.price_per_PSC
        q1 = p1.quantity
        amo1 = q1 * singlep1

        p2 = context['orders'][1]
        singlep2 = p2.price_per_PSC
        amo2 = q2 * singlep2
        q2 = p2.quantity
        

        positions = {'5': invoice_no, '17': buyer_address, '18': buyer_address, '28': p1.product.description, '31': q1,'32': singlep1, '34': amo1, '36': p2.product.description, '39': q2, '40': singlep2, '42': amo2, '44': '', '47': '', '48': '', '50': '', '52': '', '55': '', '56': '', '58': '', '60': total_psc, '61': total_amo, '66': gst, '69': gst, '72': total_psc, '73': grand_total_amo, '74': word_total_amount, '87':total_amo, '89': gst, '91': gst, '93': gst, '94': gst, '95': gst*2, '96': word_tax_total}

    elif(orders == 1):

        p1 = context['orders'][0]
        singlep1 = p1.price_per_PSC
        q1 = p1.quantity
        amo1 = q1 * singlep1

        positions = {'5': invoice_no, '17': buyer_address, '18': buyer_address, '28': p1.product.description, '31': q1, '32': singlep1, '34': amo1, '36': '', '39': '', '40': '', '42': '', '44': '', '47': '', '48': '', '50': '', '52': '', '55': '', '56': '', '58': '', '60': total_psc, '61': total_amo, '66': gst, '69': gst, '72': total_psc, '73': grand_total_amo, '74': word_total_amount, '87':total_amo, '89': gst, '91': gst, '93': gst, '94': gst, '95': gst*2, '96': word_tax_total}


    Createxl(positions)

def Createxl(context):
    cells = [5,17,18,28,31,32,34,36,39,40,42,44,47,48,50,52,55,56,58,60,61,66,69,72,73,74,87,89,91,93,94,95,96]
    wb1 = xl.load_workbook('Invoice.xlsx')
    ws1 = wb1.worksheets[0]

    counter = 0
    for i in range(ws1.max_row):
        for j in range(ws1.max_column):
            cond = ws1.cell(row = i+1, column = j+1)
            if(cond.value in cells):
                ws1.cell(row = i+1, column = j+1).value = context[f'{cells[counter]}']
                counter += 1

    wb1.save('static/new.xlsx')

def backendlogout(req):
    logout(req)
    return redirect('website:home')

class HomeView(LoginRequiredMixin,View):
    
    def get(self,request):
        clients = ClientsModel.objects.all()
        ledger = LedgerModel.objects.all()
        order = OrderModel.objects.all()
        products = ProductsModel.objects.all()
        ctx = {
            'clients':clients,
            'ledger':ledger,
            'order':order,
            'products':products,
        }
        return render(request,'backend/home.html',context=ctx)

class ProductsView(LoginRequiredMixin,View):

    def get(self,request):
        form = ProductForm()
        ctx = {
            'form':form
        }
        return render(request,'backend/products.html',context=ctx)

    def post(self,reqeust):
        form = ProductForm(reqeust.POST)
        if(form.is_valid()):
            form.save()
            return redirect('products_list')

class ProductsListView(LoginRequiredMixin,ListView):
    model = ProductsModel

class ClientsView(LoginRequiredMixin,View):
    
    def get(self,request):
        form = ClientForm()
        ctx = {
            'form':form
        }
        return render(request,'backend/clientmodel.html',context=ctx)

    def post(self,reqeust):
        form = ClientForm(reqeust.POST)
        if(form.is_valid()):
            form.save()
            return redirect('clients_list')
        else:
            return HttpResponse('in valid form')

class ClientsListView(LoginRequiredMixin,ListView):
    model = ClientsModel

class LedgerView(LoginRequiredMixin,View):
    def get(self,request):
        form = LedgerForm()
        ctx = {
            'form':form
        }
        return render(request,'backend/ledgermodel.html',context=ctx)

    def post(self,reqeust):
        form = LedgerForm(reqeust.POST)
        if(form.is_valid()):
            form.save()
            return redirect('ledger_list')
        else:
            return HttpResponse('in valid form')
        
class LedgerListView(LoginRequiredMixin,ListView):
    model = LedgerModel

class OrderListView(LoginRequiredMixin,ListView):
    model = OrderModel

class GenInvoice(View):
    def get(self,request,id):
        invoice = InvoiceModel.objects.filter(id=id)[0]
        buyer = ClientsModel.objects.filter(id = invoice.buyer_id.id)[0]
        orders = OrderModel.objects.filter(invoice_no = id)
        total_psc = 0
        total_price = 0
        for i in orders:
            total_psc += i.quantity
            total_price += i.price_per_PSC*i.quantity
        
        total_price = float(total_price)
        gst = (total_price)*(9/100)
        cs_gst = ("%.2f" % gst)
        grand_total = 2*gst+total_price
        words_grand_total = (num2words(grand_total,to='currency',lang='en_IN').replace('euro','rupees').replace('cents','paisa'))
        words_tax_total = (num2words((gst*2),to='currency',lang='en_IN').replace('euro','rupees').replace('cents','paisa'))
        words_grand_total = words_grand_total.upper()
        words_tax_total = words_tax_total.upper()
        date = datetime.datetime.now().date()
        ctx = {
            'buyer':buyer,
            'invoice':invoice,
            'orders':orders,
            'total_pices':total_psc,
            'total_amount':total_price,
            'cs_gst':cs_gst,
            'grand_total':grand_total,
            'date':date,
            'word_grand_total':words_grand_total,
            'words_tax_total':words_tax_total,
        }
        Xlgenrator(ctx)
        return render(request,'backend/output.html',context=ctx)

class FormSetView(View):

    def get(self,request):
        invoice = InvoiceForm()
        ctx = {
            'form':product,
            'invoice':invoice,
            }
        return render(request,'backend/formset.html',context=ctx)
    def post(self,request):
        data = product(request.POST)
        buyer = InvoiceForm(request.POST)
        data.is_valid() # this is necessary before makeing a object of orderform
        
        if(buyer.is_valid() and validator(data)):
            invoice_no = buyer.save()
            for i in data:
                order = OrderForm(i.cleaned_data)
                if(order.is_valid()):
                    prod = order.clean().get('product')
                    quant = order.clean().get('quantity')
                    price = order.clean().get('price_per_PSC')
                    OrderModel.objects.create(
                        product=prod,
                        quantity= quant,
                        price_per_PSC = price,
                        invoice_no = invoice_no)

            return redirect('getinvoice',id=invoice_no)

        return HttpResponse('unsuccess')